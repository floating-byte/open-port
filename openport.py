from bs4 import BeautifulSoup
import  openpyxl
from os import path
from pprint import pprint
import requests

class prox_master():
	urls = ["https://www.sslproxies.org","https://free-proxy-list.net"]
	srapper = ""

	def __init__(self,scrapper=""):
		self.scrapper = scrapper	
	def find(self,web_id):
		def web_vistor(index):
			respone = requests.get(self.urls[index])
			soup = BeautifulSoup(respone.text,"lxml")
			list_of_trs = soup.find("tbody").find_all("tr")

			json_list = []
			for i in list_of_trs:
				list_of_tds = i.find_all("td")
				ip = list_of_tds[0].text
				port = list_of_tds[1].text
				http = https = True
				if list_of_tds[5].text == "no":http=False;
				if list_of_tds[6].text == "no":https=False;
				json_list.append({"ip":ip,"port":port,"http":http,"https":https}) 
			return json_list
		try:
			self.urls[web_id-1]
			if web_id == 0: 69/0
		except Exception as e:
			raise Exception("i only support 'find(1)' and 'find(2)'")
		return web_vistor(web_id-1)	

	def setup_file(self,file_name):
		file_name = str(file_name)
		if not (".xlsx" in file_name): file_name = f"{file_name}.xlsx";
		is_exists = path.exists(file_name) 
		if not is_exists:
			wb = openpyxl.Workbook()
			ws = wb.active
			ws["A1"] = "ip"
			ws["B1"] = "port"
			ws["C1"] = "http"
			ws["D1"] = "https"
			wb.save(file_name)
	def load_data(self,file_name):
		if not (".xlsx" in file_name): file_name = f"{file_name}.xlsx";
		try:
			wb = openpyxl.load_workbook(file_name) 

		except Exception as e:
			raise Exception("the file that want to load dont exist")
	
		sheet = wb.worksheets[0]

		rows = sheet.max_row
		couls = sheet.max_column

		loaded_data = []
		for i in range(2,rows+1):
			ip = sheet[f"A{i}"].value
			port = sheet[f"B{i}"].value
			http  = sheet[f"C{i}"].value
			https = sheet[f"D{i}"].value
			loaded_data.append({"ip":ip,"port":port,"http":http,"https":https})
		return loaded_data 
	def save_data(self,file_name,list_of_data):
		if not (".xlsx" in file_name): file_name = f"{file_name}.xlsx";
		
		self.setup_file(file_name)
		wb = openpyxl.load_workbook(file_name) 
		sheet = wb.worksheets[0]
		start_row = sheet.max_row+1

		for i in range(len(list_of_data)):
			ip_pos = f"A{start_row+i}"
			port_pos = f"B{start_row+i}"
			http_pos = f"C{start_row+i}"
			https_pos = f"D{start_row+i}"
		
			sheet[ip_pos]   = list_of_data[i]["ip"]
			sheet[port_pos] = list_of_data[i]["port"]
			sheet[http_pos]   = list_of_data[i]["http"]
			sheet[https_pos] = list_of_data[i]["https"]
		
		wb.save(file_name)	
	def save(self,file_name,json_data):
		def data_checker(file_name,list_of_data):
			loaded_data = self.load_data(file_name)
			new_data = []
			for i in list_of_data:
				if not (i in loaded_data):
					new_data.append(i)
			return new_data
		file_name = f"{file_name}.xlsx"
		self.setup_file(file_name)

		new_data = data_checker(file_name,json_data)
		if len(new_data) != 0:
			self.save_data(file_name,new_data)
		print(f"{len(new_data)} are added")	


